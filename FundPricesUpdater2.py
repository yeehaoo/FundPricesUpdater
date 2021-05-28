"""
17 April 2021
github.com/yeehaoo/FundPricesUpdater

Target machine now supports python, and target website has changed, prompting a new version of the FundPricesUpdater. Old version left for posterity
Now includes stock index scraping, and automating web scraping (html file had to be manually saved in the previous version of FundPricesUpdater)

Dependencies: bs4, openpyxl, yfinance

Instructions: python3 FundPricesUpdater2.py

21 May 2021
Added calculation of changes to market indices

28 May 2021
Added rounding of calculated values to 2 decimal places
"""
import bs4
import requests
import openpyxl
import datetime
import yfinance

c = "complete"

#requesting page, parsing html and filtering prices
print("requesting page")
res = requests.get("https://www.income.com.sg/funds/fund-prices")
print(c)
print("parsing data")
soup = bs4.BeautifulSoup(res.text, "html.parser")
bid_elem = soup.select('.bid-price')
offer_elem = soup.select('.offer-price')
print(c)

#definition of lists. first, combine the prices then sort according to 
#order list
combined_list = []
sorted_list = []
correct_order_list = [14, 7, 8, 5, 1, 0, 16, 15, 3, 6, 2, 4, 9, 10, 11, 12, 19, 17, 18, 20, 13]
empty_cols = [0, 1, 2, 3, 18, 19, 30, 31, 32, 35, 36]

#combine bid and offer prices into a single list
print("combining data")
for i in range(len(bid_elem)):
	combined_list.append(str(bid_elem[i])[30:35])
	combined_list.append(str(offer_elem[i])[32:37])
print(c)

#sort the list
print("sorting list")
for i in correct_order_list:
	sorted_list.append(combined_list[i*2])
	sorted_list.append(combined_list[i*2+1])
print(c)

#adding blank values and date
print("adding filler")
for i in range(len(sorted_list)):
	if i in empty_cols:
		sorted_list.insert(i, "")
todaysDate = datetime.date.today()
todaysDate = '/'.join([str(todaysDate.day),str(todaysDate.month), str(todaysDate.year)])
sorted_list.insert(0, todaysDate)
print(c)

#stock indexes
print("requesting stock data")
stock_indexes = ["^STI", "^IXIC", "^GSPC", "^DJI"]
for index in stock_indexes:
	stock_value = yfinance.Ticker(index).info["regularMarketPrice"]
	sorted_list.append(str(stock_value))
	sorted_list.append("")
	print(index, stock_value)
print(c)

#load workbook (ws stands for worksheet)
print("loading workbook")
workbook = openpyxl.load_workbook(filename = 'fund.xlsx')
ws = workbook.active
row_count = ws.max_row + 1
print(c)

#write values
print("writing values")
pause_columns = [55, 57, 59, 61]
for i in range(len(sorted_list)):
	if i in pause_columns:
		value_today = float(sorted_list[i-1])
		value_yesterday = float(ws.cell(row=row_count-1, column=i).value)
		calculated_difference = value_today - value_yesterday
		ws.cell(row=row_count, column=i+1, value=str(round(calculated_difference,2)))
	else:
		ws.cell(row=row_count, column=i+1, value=sorted_list[i])

print(c)

#save
print("saving workbook")
workbook.save('fundwritten.xlsx')
print(c)
